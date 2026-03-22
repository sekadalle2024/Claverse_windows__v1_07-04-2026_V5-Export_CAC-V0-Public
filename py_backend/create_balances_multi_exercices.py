# -*- coding: utf-8 -*-
"""
Script pour créer un fichier Excel avec 3 balances (N, N-1, N-2)
et examiner la structure de la liasse officielle SYSCOHADA
"""
import pandas as pd
import openpyxl
import random
import sys

# Forcer l'encodage UTF-8 pour la sortie
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Charger la balance démo existante
balance_demo = pd.read_excel('P000 -BALANCE DEMO.xls')
print(f"Balance demo chargee: {len(balance_demo)} lignes")
print(f"Colonnes: {balance_demo.columns.tolist()}")

# Créer 3 versions de la balance avec variations
def create_balance_variation(base_df, year_offset, variation_pct=0.15):
    """Crée une variation de la balance pour un exercice différent"""
    df = base_df.copy()
    
    # Appliquer des variations aléatoires sur les montants
    for col in df.columns:
        if 'solde' in col.lower() or 'debit' in col.lower() or 'credit' in col.lower():
            if df[col].dtype in ['float64', 'int64']:
                # Variation aléatoire entre -variation_pct et +variation_pct
                variation = df[col] * (1 + random.uniform(-variation_pct, variation_pct))
                df[col] = variation.round(2)
    
    return df

# Créer les 3 balances
balance_n = balance_demo.copy()
balance_n1 = create_balance_variation(balance_demo, -1, 0.10)  # -10% à +10%
balance_n2 = create_balance_variation(balance_demo, -2, 0.20)  # -20% à +20%

# Créer le fichier Excel avec 3 onglets
output_file = 'BALANCES_N_N1_N2.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    balance_n.to_excel(writer, sheet_name='Balance N (2024)', index=False)
    balance_n1.to_excel(writer, sheet_name='Balance N-1 (2023)', index=False)
    balance_n2.to_excel(writer, sheet_name='Balance N-2 (2022)', index=False)

print(f"\nFichier cree: {output_file}")
print(f"   - Onglet 1: Balance N (2024) - {len(balance_n)} comptes")
print(f"   - Onglet 2: Balance N-1 (2023) - {len(balance_n1)} comptes")
print(f"   - Onglet 3: Balance N-2 (2022) - {len(balance_n2)} comptes")

# Examiner la structure de la liasse officielle
print("\n" + "="*80)
print("EXAMEN DE LA LIASSE OFFICIELLE")
print("="*80)

try:
    wb = openpyxl.load_workbook('Liasse officielle.xlsm', data_only=True)
    print(f"\nOnglets disponibles: {wb.sheetnames}")
    
    # Examiner chaque onglet pertinent
    for sheet_name in wb.sheetnames:
        if any(keyword in sheet_name.lower() for keyword in ['bilan', 'actif', 'passif', 'resultat', 'compte']):
            ws = wb[sheet_name]
            print(f"\n{'='*60}")
            print(f"ONGLET: {sheet_name}")
            print(f"{'='*60}")
            print(f"Dimensions: {ws.max_row} lignes x {ws.max_column} colonnes")
            
            # Afficher les 20 premières lignes
            print(f"\nPremieres lignes:")
            for row_idx in range(1, min(21, ws.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(10, ws.max_column + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    if value:
                        row_data.append(f"[{col_idx}] {str(value)[:40]}")
                
                if row_data:
                    print(f"Ligne {row_idx}: {' | '.join(row_data)}")
    
    wb.close()
    
except Exception as e:
    print(f"Erreur lors de l'examen de la liasse: {e}")

print("\nScript termine")
